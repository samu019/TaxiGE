from datetime import date
from django.core.paginator import Paginator
from notifications.models import Notification
from audits.models import AuditLog
from sharing.models import CompanyMember, TaxiShareInvite
from django.contrib import messages
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import csv
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse

from companies.models import Company
from vehicles.models import Vehicle
from drivers.models import Driver
from payments.models import DriverPayment
from damages.models import VehicleDamage
from .forms import VehicleForm, DriverForm, DriverPaymentForm, VehicleDamageForm
from django.db.models import Q, Sum


def user_can_access_panel(user):
    return bool(
        user.is_authenticated
        and user.is_active
        and user.role == 'admin'
        and user.admin_approved
    )




def subscription_guard(request):
    if not request.user.is_authenticated:
        return None

    if request.user.is_superuser or getattr(request.user, "is_platform_owner", False):
        return None

    try:
        if request.user.saas_subscription.is_valid():
            return None
    except Exception:
        pass

    messages.error(
        request,
        "Tu suscripción SaaS no está activa. Renueva tu suscripción para acceder al panel."
    )
    return redirect("/panel/subscription/renew/")


def panel_guard(request):
    user = request.user

    if user.is_superuser and user.is_platform_owner:
        return redirect('/admin/')

    if not user_can_access_panel(user):
        return HttpResponseForbidden(
            'Acceso denegado. Tu cuenta aún no está aprobada como dueño de taxi.'
        )

    return None


def get_user_company(user):
    return Company.objects.filter(owner=user, status='active').first()


def calculate_payment_status(expected_amount, paid_amount):
    if paid_amount <= 0:
        return 'pending'

    if paid_amount >= expected_amount:
        return 'paid'

    return 'partial'







def get_client_ip(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')






def render_panel(request, template_name, context=None):
    context = context or {}
    try:
        context['global_unread_notifications'] = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
    except Exception:
        context['global_unread_notifications'] = 0

    return render(request, template_name, context)


def create_notification(user, title, message, level='info', link=None):
    try:
        recent_exists = Notification.objects.filter(
            user=user,
            title=title,
            message=message,
            is_read=False,
        ).exists()

        if recent_exists:
            return None

        return Notification.objects.create(
            user=user,
            title=title,
            message=message,
            level=level,
            link=link,
        )
    except Exception:
        return None


def audit_event(request, action, app_label, model_name, object_id=None, object_repr=None, description=None):
    try:
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            app_label=app_label,
            model_name=model_name,
            object_id=str(object_id) if object_id is not None else None,
            object_repr=str(object_repr)[:255] if object_repr else None,
            description=description,
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
        )
    except Exception:
        pass


def shared_permissions_for_role(role):
    permissions = {
        CompanyMember.ROLE_PARTNER: {
            'can_view_payments': True,
            'can_view_damages': True,
            'can_add_payments': False,
            'can_add_damages': False,
            'can_edit_data': False,
            'can_invite_members': False,
            'can_export_reports': True,
        },
        CompanyMember.ROLE_MANAGER: {
            'can_view_payments': True,
            'can_view_damages': True,
            'can_add_payments': True,
            'can_add_damages': True,
            'can_edit_data': True,
            'can_invite_members': False,
            'can_export_reports': False,
        },
        CompanyMember.ROLE_VIEWER: {
            'can_view_payments': True,
            'can_view_damages': True,
            'can_add_payments': False,
            'can_add_damages': False,
            'can_edit_data': False,
            'can_invite_members': False,
            'can_export_reports': False,
        },
        CompanyMember.ROLE_ACCOUNTANT: {
            'can_view_payments': True,
            'can_view_damages': False,
            'can_add_payments': False,
            'can_add_damages': False,
            'can_edit_data': False,
            'can_invite_members': False,
            'can_export_reports': True,
        },
    }
    return permissions.get(role, permissions[CompanyMember.ROLE_VIEWER])




def user_is_company_owner(user, company):
    return company.owner_id == user.id


def get_company_membership(user, company):
    return CompanyMember.objects.filter(
        user=user,
        company=company,
        is_active=True
    ).first()


def user_can_add_payments(user, company):
    if user_is_company_owner(user, company):
        return True
    member = get_company_membership(user, company)
    return bool(member and member.can_add_payments)


def user_can_add_damages(user, company):
    if user_is_company_owner(user, company):
        return True
    member = get_company_membership(user, company)
    return bool(member and member.can_add_damages)


def user_can_edit_data(user, company):
    if user_is_company_owner(user, company):
        return True
    member = get_company_membership(user, company)
    return bool(member and member.can_edit_data)




def user_has_any_export_permission(user):
    companies = Company.objects.filter(user_company_direct_access_q(user)).distinct()
    for company in companies:
        if user_can_export_reports(user, company):
            return True
    return False


def user_can_export_reports(user, company):
    if user_is_company_owner(user, company):
        return True
    member = get_company_membership(user, company)
    return bool(member and member.can_export_reports)


def user_company_access_q(user):
    return (
        Q(company__owner=user) |
        Q(company__members__user=user, company__members__is_active=True)
    )


def user_company_direct_access_q(user):
    return (
        Q(owner=user) |
        Q(members__user=user, members__is_active=True)
    )

@login_required
def panel_home(request):
    from django.utils import timezone

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    user = request.user

    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)
    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)

    today = date.today()
    month_start = today.replace(day=1)
    last_7_start = today - timezone.timedelta(days=7)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid

    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost

    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    month_payments = payments.filter(payment_date__gte=month_start, payment_date__lte=today)
    today_payments = payments.filter(payment_date=today)
    last_7_payments = payments.filter(payment_date__gte=last_7_start, payment_date__lte=today)
    month_damages = damages.filter(damage_date__gte=month_start, damage_date__lte=today)

    month_expected = month_payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    month_paid = month_payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    month_debt = month_expected - month_paid
    month_damages_cost = month_damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    month_net_profit = month_paid - month_damages_cost

    today_paid = today_payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    last_7_paid = last_7_payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')

    active_vehicles = vehicles.filter(status='active').count()
    active_drivers = drivers.filter(status='active').count()
    damages_pending = damages.filter(status='pending').count()

    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    context = {
        'companies_count': companies.count(),
        'vehicles_count': vehicles.count(),
        'drivers_count': drivers.count(),
        'payments_count': payments.count(),
        'damages_count': damages.count(),

        'active_vehicles': active_vehicles,
        'active_drivers': active_drivers,
        'damages_pending': damages_pending,

        'total_expected': total_expected,
        'total_paid': total_paid,
        'total_debt': total_debt,
        'damages_cost': damages_cost,
        'net_profit': net_profit,
        'collection_rate': collection_rate,

        'month_label': today.strftime('%m/%Y'),
        'month_expected': month_expected,
        'month_paid': month_paid,
        'month_debt': month_debt,
        'month_damages_cost': month_damages_cost,
        'month_net_profit': month_net_profit,
        'today_paid': today_paid,
        'last_7_paid': last_7_paid,

        'recent_payments': payments.order_by('-payment_date', '-created_at')[:5],
        'recent_damages': damages.order_by('-damage_date', '-created_at')[:5],

        'active_filter': filter_type,
        'start_date': start_date or '',
        'end_date': end_date or '',

        'chart_paid': float(total_paid),
        'chart_debt': float(total_debt),
        'chart_damages': float(damages_cost),
        'chart_net': float(net_profit),
        'chart_active_vehicles': active_vehicles,
        'chart_active_drivers': active_drivers,
        'chart_damages_pending': damages_pending,
        'chart_payments_paid': payments.filter(status='paid').count(),
        'chart_payments_partial': payments.filter(status='partial').count(),
        'chart_payments_pending': payments.filter(status='pending').count(),
    }

    return render_panel(request, 'dashboard/panel_home.html', context)



@login_required
def panel_vehicles(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        vehicles = vehicles.filter(
            Q(plate_number__icontains=q) |
            Q(brand__icontains=q) |
            Q(model__icontains=q) |
            Q(color__icontains=q)
        )

    if status:
        vehicles = vehicles.filter(status=status).select_related('company')

    paginator = Paginator(vehicles, 10)
    page_number = request.GET.get('page')
    vehicles_page = paginator.get_page(page_number)

    return render_panel(request, 'dashboard/panel_vehicles.html', {
        'vehicles': vehicles_page,
    })


@login_required
def panel_vehicle_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)

        if form.is_valid():
            vehicle = form.save(commit=False)
            vehicle.company = company
            vehicle.created_by = request.user
            vehicle.save()
            return redirect('/panel/vehicles/')
    else:
        form = VehicleForm()

    return render(request, 'dashboard/panel_vehicle_form.html', {
        'form': form,
        'title': 'Registrar nuevo taxi',
        'button_text': 'Guardar taxi',
    })


@login_required
def panel_vehicle_edit(request, vehicle_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    vehicle = get_object_or_404(
        Vehicle.objects.filter(user_company_access_q(request.user)),
        id=vehicle_id
    )

    if request.method == 'POST':
        if not user_can_edit_data(request.user, vehicle.company):
            messages.error(request, 'No tienes permiso para editar este taxi.')
            return redirect('/panel/vehicles/')
        form = VehicleForm(request.POST, request.FILES, instance=vehicle)

        if form.is_valid():
            form.save()
            return redirect('/panel/vehicles/')
    else:
        form = VehicleForm(instance=vehicle)

    return render(request, 'dashboard/panel_vehicle_form.html', {
        'form': form,
        'title': 'Editar taxi',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_drivers(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    drivers = Driver.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        drivers = drivers.filter(
            Q(full_name__icontains=q) |
            Q(phone__icontains=q) |
            Q(identity_number__icontains=q) |
            Q(license_number__icontains=q) |
            Q(assigned_vehicle__plate_number__icontains=q)
        )

    if status:
        drivers = drivers.filter(status=status).select_related('company', 'assigned_vehicle')

    paginator = Paginator(drivers, 10)
    page_number = request.GET.get('page')
    drivers_page = paginator.get_page(page_number)

    return render_panel(request, 'dashboard/panel_drivers.html', {
        'drivers': drivers_page,
    })


@login_required
def panel_driver_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        form = DriverForm(request.POST, request.FILES, company=company)

        if form.is_valid():
            driver = form.save(commit=False)
            driver.company = company
            driver.created_by = request.user
            driver.save()
            return redirect('/panel/drivers/')
    else:
        form = DriverForm(company=company)

    return render(request, 'dashboard/panel_driver_form.html', {
        'form': form,
        'title': 'Registrar nuevo conductor',
        'button_text': 'Guardar conductor',
    })


@login_required
def panel_driver_edit(request, driver_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    company = get_user_company(request.user)

    driver = get_object_or_404(
        Driver.objects.filter(user_company_access_q(request.user)),
        id=driver_id
    )

    if request.method == 'POST':
        if not user_can_edit_data(request.user, driver.company):
            messages.error(request, 'No tienes permiso para editar este conductor.')
            return redirect('/panel/drivers/')
        form = DriverForm(request.POST, request.FILES, instance=driver, company=company)

        if form.is_valid():
            form.save()
            return redirect('/panel/drivers/')
    else:
        form = DriverForm(instance=driver, company=company)

    return render(request, 'dashboard/panel_driver_form.html', {
        'form': form,
        'title': 'Editar conductor',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_payments(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    payments = DriverPayment.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        payments = payments.filter(
            Q(driver__full_name__icontains=q) |
            Q(vehicle__plate_number__icontains=q) |
            Q(notes__icontains=q)
        )

    if status:
        payments = payments.filter(status=status).select_related(
        'company',
        'driver',
        'vehicle',
    ).order_by('-payment_date', '-created_at')

    paginator = Paginator(payments, 10)
    page_number = request.GET.get('page')
    payments_page = paginator.get_page(page_number)

    return render_panel(request, 'dashboard/panel_payments.html', {
        'payments': payments_page,
    })


@login_required
def panel_payment_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        if not user_can_add_payments(request.user, company):
            messages.error(request, 'No tienes permiso para registrar pagos.')
            return redirect('/panel/payments/')
        form = DriverPaymentForm(request.POST, company=company)

        if form.is_valid():
            payment = form.save(commit=False)
            payment.company = company
            payment.registered_by = request.user
            payment.status = calculate_payment_status(
                payment.expected_amount,
                payment.paid_amount
            )
            payment.save()

            if payment.debt_amount > 0:
                create_notification(
                    payment.company.owner,
                    'Deuda pendiente registrada',
                    f'El conductor {payment.driver} tiene una deuda pendiente de {payment.debt_amount} XAF.',
                    'warning',
                    '/panel/payments/'
                )

            return redirect('/panel/payments/')
    else:
        form = DriverPaymentForm(company=company)

    return render(request, 'dashboard/panel_payment_form.html', {
        'form': form,
        'title': 'Registrar pago',
        'button_text': 'Guardar pago',
    })

@login_required
def panel_payment_edit(request, payment_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    company = get_user_company(request.user)

    payment = get_object_or_404(
        DriverPayment.objects.filter(user_company_access_q(request.user)),
        id=payment_id
    )

    if request.method == 'POST':
        if not user_can_edit_data(request.user, payment.company):
            messages.error(request, 'No tienes permiso para editar este pago.')
            return redirect('/panel/payments/')
        form = DriverPaymentForm(request.POST, instance=payment, company=company)

        if form.is_valid():
            payment = form.save(commit=False)
            payment.company = company
            payment.registered_by = request.user
            payment.status = calculate_payment_status(
                payment.expected_amount,
                payment.paid_amount
            )
            payment.save()

            if payment.debt_amount > 0:
                create_notification(
                    payment.company.owner,
                    'Deuda pendiente registrada',
                    f'El conductor {payment.driver} tiene una deuda pendiente de {payment.debt_amount} XAF.',
                    'warning',
                    '/panel/payments/'
                )

            return redirect('/panel/payments/')
    else:
        form = DriverPaymentForm(instance=payment, company=company)

    return render(request, 'dashboard/panel_payment_form.html', {
        'form': form,
        'title': 'Editar pago',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_damages(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        damages = damages.filter(
            Q(title__icontains=q) |
            Q(description__icontains=q) |
            Q(vehicle__plate_number__icontains=q) |
            Q(driver__full_name__icontains=q)
        )

    if status:
        damages = damages.filter(status=status).select_related(
        'company',
        'driver',
        'vehicle',
    ).order_by('-damage_date', '-created_at')

    paginator = Paginator(damages, 10)
    page_number = request.GET.get('page')
    damages_page = paginator.get_page(page_number)

    return render_panel(request, 'dashboard/panel_damages.html', {
        'damages': damages_page,
    })


@login_required
def panel_damage_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        if not user_can_add_damages(request.user, company):
            messages.error(request, 'No tienes permiso para registrar daños.')
            return redirect('/panel/damages/')
        form = VehicleDamageForm(request.POST, request.FILES, company=company)

        if form.is_valid():
            damage = form.save(commit=False)
            damage.company = company
            damage.registered_by = request.user
            damage.save()

            create_notification(
                damage.company.owner,
                'Nuevo daño registrado',
                f'Se registró un daño en {damage.vehicle}: {damage.title}. Costo estimado: {damage.estimated_cost} XAF.',
                'danger' if damage.status == 'pending' else 'warning',
                '/panel/damages/'
            )

            return redirect('/panel/damages/')
    else:
        form = VehicleDamageForm(company=company)

    return render(request, 'dashboard/panel_damage_form.html', {
        'form': form,
        'title': 'Registrar daño',
        'button_text': 'Guardar daño',
    })

@login_required
def panel_damage_edit(request, damage_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    company = get_user_company(request.user)

    damage = get_object_or_404(
        VehicleDamage.objects.filter(user_company_access_q(request.user)),
        id=damage_id
    )

    if request.method == 'POST':
        if not user_can_edit_data(request.user, damage.company):
            messages.error(request, 'No tienes permiso para editar este daño.')
            return redirect('/panel/damages/')
        form = VehicleDamageForm(request.POST, request.FILES, instance=damage, company=company)

        if form.is_valid():
            damage = form.save(commit=False)
            damage.company = company
            damage.registered_by = request.user
            damage.save()

            create_notification(
                damage.company.owner,
                'Nuevo daño registrado',
                f'Se registró un daño en {damage.vehicle}: {damage.title}. Costo estimado: {damage.estimated_cost} XAF.',
                'danger' if damage.status == 'pending' else 'warning',
                '/panel/damages/'
            )

            return redirect('/panel/damages/')
    else:
        form = VehicleDamageForm(instance=damage, company=company)

    return render(request, 'dashboard/panel_damage_form.html', {
        'form': form,
        'title': 'Editar daño',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_vehicle_delete(request, vehicle_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    vehicle = get_object_or_404(Vehicle.objects.filter(user_company_access_q(request.user)), id=vehicle_id)

    if request.method == 'POST':
        if not user_can_edit_data(request.user, vehicle.company):
            messages.error(request, 'No tienes permiso para eliminar este taxi.')
            return redirect('/panel/vehicles/')
        vehicle.delete()
        return redirect('/panel/vehicles/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar taxi',
        'object_name': vehicle.plate_number,
        'cancel_url': '/panel/vehicles/',
    })


@login_required
def panel_driver_delete(request, driver_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    driver = get_object_or_404(Driver.objects.filter(user_company_access_q(request.user)), id=driver_id)

    if request.method == 'POST':
        if not user_can_edit_data(request.user, driver.company):
            messages.error(request, 'No tienes permiso para eliminar este conductor.')
            return redirect('/panel/drivers/')
        driver.delete()
        return redirect('/panel/drivers/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar conductor',
        'object_name': driver.full_name,
        'cancel_url': '/panel/drivers/',
    })


@login_required
def panel_payment_delete(request, payment_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    payment = get_object_or_404(DriverPayment.objects.filter(user_company_access_q(request.user)), id=payment_id)

    if request.method == 'POST':
        if not user_can_edit_data(request.user, payment.company):
            messages.error(request, 'No tienes permiso para eliminar este pago.')
            return redirect('/panel/payments/')
        payment.delete()
        return redirect('/panel/payments/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar pago',
        'object_name': f'{payment.driver.full_name} - {payment.payment_date}',
        'cancel_url': '/panel/payments/',
    })


@login_required
def panel_damage_delete(request, damage_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    damage = get_object_or_404(VehicleDamage.objects.filter(user_company_access_q(request.user)), id=damage_id)

    if request.method == 'POST':
        if not user_can_edit_data(request.user, damage.company):
            messages.error(request, 'No tienes permiso para eliminar este daño.')
            return redirect('/panel/damages/')
        damage.delete()
        return redirect('/panel/damages/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar daño',
        'object_name': damage.title,
        'cancel_url': '/panel/damages/',
    })


@login_required
def api_panel_summary(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    user = request.user

    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)
    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or 0
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or 0
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost

    if total_expected > 0:
        collection_rate = round((total_paid / total_expected) * 100, 2)
    else:
        collection_rate = Decimal('0')
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or 0

    data = {
        'companies_count': companies.count(),
        'vehicles_count': vehicles.count(),
        'active_vehicles': vehicles.filter(status='active').count(),
        'drivers_count': drivers.count(),
        'active_drivers': drivers.filter(status='active').count(),
        'payments_count': payments.count(),
        'damages_count': damages.count(),
        'damages_pending': damages.filter(status='pending').count(),
        'total_expected': str(total_expected),
        'total_paid': str(total_paid),
        'total_debt': str(total_debt),
        'damages_cost': str(damages_cost),
    }

    return JsonResponse({
        'ok': True,
        'data': data,
    })


@login_required
def api_panel_vehicles(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user)).select_related('company').order_by('-created_at')

    data = []
    for v in vehicles:
        data.append({
            'id': v.id,
            'plate_number': v.plate_number,
            'brand': v.brand,
            'model': v.model,
            'color': v.color,
            'year': v.year,
            'daily_target_amount': str(v.daily_target_amount),
            'status': v.status,
            'status_display': v.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


@login_required
def api_panel_drivers(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    drivers = Driver.objects.filter(user_company_access_q(request.user)).select_related('assigned_vehicle').order_by('-created_at')

    data = []
    for d in drivers:
        data.append({
            'id': d.id,
            'full_name': d.full_name,
            'phone': d.phone,
            'assigned_vehicle': d.assigned_vehicle.plate_number if d.assigned_vehicle else None,
            'daily_payment_amount': str(d.daily_payment_amount),
            'payment_day': d.payment_day,
            'status': d.status,
            'status_display': d.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


@login_required
def api_panel_payments(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    payments = DriverPayment.objects.filter(user_company_access_q(request.user)).select_related('driver', 'vehicle').order_by('-payment_date', '-created_at')[:100]

    data = []
    for p in payments:
        data.append({
            'id': p.id,
            'driver': p.driver.full_name,
            'vehicle': p.vehicle.plate_number if p.vehicle else None,
            'payment_date': str(p.payment_date),
            'expected_amount': str(p.expected_amount),
            'paid_amount': str(p.paid_amount),
            'debt_amount': str(p.debt_amount),
            'status': p.status,
            'status_display': p.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


@login_required
def api_panel_damages(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user)).select_related('driver', 'vehicle').order_by('-damage_date', '-created_at')[:100]

    data = []
    for d in damages:
        data.append({
            'id': d.id,
            'vehicle': d.vehicle.plate_number if d.vehicle else None,
            'driver': d.driver.full_name if d.driver else None,
            'title': d.title,
            'damage_date': str(d.damage_date),
            'estimated_cost': str(d.estimated_cost),
            'final_cost': str(d.final_cost),
            'status': d.status,
            'status_display': d.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


def make_csv_response(filename, headers, rows):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(headers)

    for row in rows:
        writer.writerow(row)

    return response


@login_required
def export_vehicles_csv(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user)).select_related('company').order_by('-created_at')

    rows = []
    for v in vehicles:
        rows.append([
            v.id,
            v.company.name,
            v.plate_number,
            v.internal_code,
            v.brand,
            v.model,
            v.color,
            v.year,
            v.daily_target_amount,
            v.get_status_display(),
            v.created_at,
        ])

    return make_csv_response(
        'taxige_taxis.csv',
        ['ID', 'Empresa', 'Matrícula', 'Código interno', 'Marca', 'Modelo', 'Color', 'Año', 'Objetivo diario', 'Estado', 'Creado'],
        rows
    )


@login_required
def export_drivers_csv(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    drivers = Driver.objects.filter(user_company_access_q(request.user)).select_related('company', 'assigned_vehicle').order_by('-created_at')

    rows = []
    for d in drivers:
        rows.append([
            d.id,
            d.company.name,
            d.full_name,
            d.phone,
            d.identity_number,
            d.license_number,
            d.assigned_vehicle.plate_number if d.assigned_vehicle else '',
            d.daily_payment_amount,
            d.payment_day,
            d.get_status_display(),
            d.created_at,
        ])

    return make_csv_response(
        'taxige_conductores.csv',
        ['ID', 'Empresa', 'Conductor', 'Teléfono', 'Documento', 'Licencia', 'Taxi asignado', 'Pago diario', 'Día de pago', 'Estado', 'Creado'],
        rows
    )


@login_required
def export_payments_csv(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    payments = DriverPayment.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-payment_date', '-created_at')

    rows = []
    for p in payments:
        rows.append([
            p.id,
            p.company.name,
            p.driver.full_name,
            p.vehicle.plate_number if p.vehicle else '',
            p.payment_date,
            p.expected_amount,
            p.paid_amount,
            p.debt_amount,
            p.get_status_display(),
            p.notes,
            p.created_at,
        ])

    return make_csv_response(
        'taxige_pagos.csv',
        ['ID', 'Empresa', 'Conductor', 'Taxi', 'Fecha', 'Esperado', 'Pagado', 'Deuda', 'Estado', 'Notas', 'Creado'],
        rows
    )


@login_required
def export_damages_csv(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-damage_date', '-created_at')

    rows = []
    for d in damages:
        rows.append([
            d.id,
            d.company.name,
            d.vehicle.plate_number if d.vehicle else '',
            d.driver.full_name if d.driver else '',
            d.title,
            d.damage_date,
            d.estimated_cost,
            d.final_cost,
            d.get_status_display(),
            d.description,
            d.created_at,
        ])

    return make_csv_response(
        'taxige_danos.csv',
        ['ID', 'Empresa', 'Taxi', 'Conductor', 'Daño', 'Fecha', 'Costo estimado', 'Costo final', 'Estado', 'Descripción', 'Creado'],
        rows
    )



def make_pdf_response(filename, title, headers, rows):
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak
    from django.utils import timezone

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    generated_at = timezone.localtime().strftime('%d/%m/%Y %H:%M')

    header_data = [
        [
            Paragraph('<b>TaxiGE Platform</b><br/><font size="8">Sistema profesional de gestión de taxis</font>', styles['Normal']),
            Paragraph(f'<b>{title}</b><br/><font size="8">Generado: {generated_at}</font>', styles['Normal']),
        ]
    ]

    header_table = Table(header_data, colWidths=[130 * mm, 130 * mm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#111827')),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#334155')),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 14))

    summary_data = [
        ['Total de registros', str(len(rows)), 'Formato', 'PDF corporativo'],
    ]

    summary_table = Table(summary_data, colWidths=[45 * mm, 35 * mm, 35 * mm, 55 * mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0F172A')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 14))

    safe_rows = rows if rows else [['Sin datos'] + ['' for _ in headers[1:]]]
    data = [headers] + safe_rows

    available_width = landscape(A4)[0] - (32 * mm)
    col_count = len(headers)
    col_widths = [available_width / col_count for _ in headers]

    table = Table(data, repeatRows=1, colWidths=col_widths)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        '<font size="8" color="#64748B">Documento generado automáticamente por TaxiGE Platform. Uso interno y administrativo.</font>',
        styles['Normal']
    ))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#64748B'))
        canvas.drawString(16 * mm, 10 * mm, 'TaxiGE Platform')
        canvas.drawRightString(281 * mm, 10 * mm, f'Página {doc.page}')
        canvas.restoreState()

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)

    return response


@login_required
def export_vehicles_pdf(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user)).select_related('company').order_by('-created_at')

    rows = [[
        str(v.id),
        v.company.name,
        v.plate_number,
        v.brand,
        v.model,
        v.color,
        str(v.daily_target_amount),
        v.get_status_display(),
    ] for v in vehicles]

    return make_pdf_response(
        'taxige_taxis.pdf',
        'Reporte de taxis',
        ['ID', 'Empresa', 'Matrícula', 'Marca', 'Modelo', 'Color', 'Objetivo diario', 'Estado'],
        rows
    )


@login_required
def export_drivers_pdf(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    drivers = Driver.objects.filter(user_company_access_q(request.user)).select_related('company', 'assigned_vehicle').order_by('-created_at')

    rows = [[
        str(d.id),
        d.company.name,
        d.full_name,
        d.phone,
        d.assigned_vehicle.plate_number if d.assigned_vehicle else '',
        str(d.daily_payment_amount),
        d.payment_day,
        d.get_status_display(),
    ] for d in drivers]

    return make_pdf_response(
        'taxige_conductores.pdf',
        'Reporte de conductores',
        ['ID', 'Empresa', 'Conductor', 'Teléfono', 'Taxi', 'Pago diario', 'Día pago', 'Estado'],
        rows
    )


@login_required
def export_payments_pdf(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    payments = DriverPayment.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-payment_date', '-created_at')

    rows = [[
        str(p.id),
        p.driver.full_name,
        p.vehicle.plate_number if p.vehicle else '',
        str(p.payment_date),
        str(p.expected_amount),
        str(p.paid_amount),
        str(p.debt_amount),
        p.get_status_display(),
    ] for p in payments]

    return make_pdf_response(
        'taxige_pagos.pdf',
        'Reporte de pagos',
        ['ID', 'Conductor', 'Taxi', 'Fecha', 'Esperado', 'Pagado', 'Deuda', 'Estado'],
        rows
    )


@login_required
def export_damages_pdf(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-damage_date', '-created_at')

    rows = [[
        str(d.id),
        d.vehicle.plate_number if d.vehicle else '',
        d.driver.full_name if d.driver else '',
        d.title,
        str(d.damage_date),
        str(d.estimated_cost),
        str(d.final_cost),
        d.get_status_display(),
    ] for d in damages]

    return make_pdf_response(
        'taxige_danos.pdf',
        'Reporte de daños',
        ['ID', 'Taxi', 'Conductor', 'Daño', 'Fecha', 'Estimado', 'Final', 'Estado'],
        rows
    )


@login_required
def dashboard_export_excel(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    from django.http import HttpResponse
    from django.db.models import Sum
    from decimal import Decimal
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from payments.models import DriverPayment
    from damages.models import VehicleDamage
    from companies.models import Company
    from vehicles.models import Vehicle
    from drivers.models import Driver

    user = request.user

    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)
    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost
    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    wb = Workbook()

    dark = "0F172A"
    yellow = "FACC15"
    green = "22C55E"
    gray = "E2E8F0"

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ======================
    # HOJA RESUMEN
    # ======================
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "Reporte Ejecutivo TaxiGE"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Usuario: {user.username}"
    ws["A3"] = "Sistema: TaxiGE Platform"

    resumen = [
        ["Indicador", "Valor"],
        ["Ingresos cobrados", float(total_paid)],
        ["Deuda pendiente", float(total_debt)],
        ["Costo daños", float(damages_cost)],
        ["Beneficio neto", float(net_profit)],
        ["Ratio de cobro", f"{collection_rate}%"],
        ["Empresas", companies.count()],
        ["Taxis activos", vehicles.filter(status='active').count()],
        ["Conductores activos", drivers.filter(status='active').count()],
        ["Daños pendientes", damages.filter(status='pending').count()],
    ]

    start_row = 5
    for r, row in enumerate(resumen, start=start_row):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r == start_row:
                cell.font = Font(bold=True, color="111827")
                cell.fill = PatternFill("solid", fgColor=yellow)

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 26

    # ======================
    # HOJA PAGOS
    # ======================
    ws2 = wb.create_sheet("Pagos")
    headers = ["Conductor", "Taxi", "Fecha", "Esperado", "Pagado", "Deuda", "Estado"]

    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r, pay in enumerate(payments.order_by("-payment_date", "-created_at"), 2):
        row = [
            getattr(pay.driver, "full_name", ""),
            getattr(pay.vehicle, "plate_number", ""),
            str(pay.payment_date),
            float(pay.expected_amount),
            float(pay.paid_amount),
            float(pay.debt_amount),
            pay.get_status_display(),
        ]
        for c, value in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ======================
    # HOJA DAÑOS
    # ======================
    ws3 = wb.create_sheet("Daños")
    headers = ["Taxi", "Conductor", "Daño", "Fecha", "Costo estimado", "Estado"]

    for c, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r, dmg in enumerate(damages.order_by("-damage_date", "-created_at"), 2):
        row = [
            getattr(dmg.vehicle, "plate_number", ""),
            getattr(dmg.driver, "full_name", ""),
            dmg.title,
            str(dmg.damage_date),
            float(dmg.estimated_cost),
            dmg.get_status_display(),
        ]
        for c, value in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 24

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="taxige_reporte_ejecutivo.xlsx"'
    wb.save(response)
    return response


@login_required
def dashboard_export_pdf(request):
    if not user_has_any_export_permission(request.user):
        audit_event(
            request,
            AuditLog.ACTION_VIEW,
            'dashboard',
            'Export',
            None,
            'Exportación bloqueada',
            'Intento de exportación sin permiso'
        )
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('/panel/')

    from django.http import HttpResponse
    from django.db.models import Sum
    from decimal import Decimal
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from payments.models import DriverPayment
    from damages.models import VehicleDamage
    from companies.models import Company
    from vehicles.models import Vehicle
    from drivers.models import Driver

    user = request.user

    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)
    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost
    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="taxige_reporte_ejecutivo.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("<b>Reporte Ejecutivo TaxiGE</b>", styles["Title"]))
    story.append(Paragraph(f"Usuario: {user.username}", styles["Normal"]))
    story.append(Spacer(1, 14))

    resumen = [
        ["Indicador", "Valor"],
        ["Ingresos cobrados", f"{total_paid} XAF"],
        ["Deuda pendiente", f"{total_debt} XAF"],
        ["Costo daños", f"{damages_cost} XAF"],
        ["Beneficio neto", f"{net_profit} XAF"],
        ["Ratio de cobro", f"{collection_rate}%"],
        ["Empresas", companies.count()],
        ["Taxis activos", vehicles.filter(status='active').count()],
        ["Conductores activos", drivers.filter(status='active').count()],
        ["Daños pendientes", damages.filter(status='pending').count()],
    ]

    table = Table(resumen, colWidths=[220, 220])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Últimos pagos</b>", styles["Heading2"]))
    pagos_data = [["Conductor", "Taxi", "Fecha", "Pagado", "Estado"]]
    for pay in payments.order_by("-payment_date", "-created_at")[:8]:
        pagos_data.append([
            getattr(pay.driver, "full_name", ""),
            getattr(pay.vehicle, "plate_number", ""),
            str(pay.payment_date),
            f"{pay.paid_amount} XAF",
            pay.get_status_display(),
        ])

    pagos_table = Table(pagos_data, colWidths=[120, 80, 85, 95, 95])
    pagos_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#facc15")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(pagos_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Últimos daños</b>", styles["Heading2"]))
    damages_data = [["Taxi", "Conductor", "Daño", "Costo", "Estado"]]
    for dmg in damages.order_by("-damage_date", "-created_at")[:8]:
        damages_data.append([
            getattr(dmg.vehicle, "plate_number", ""),
            getattr(dmg.driver, "full_name", ""),
            dmg.title,
            f"{dmg.estimated_cost} XAF",
            dmg.get_status_display(),
        ])

    damages_table = Table(damages_data, colWidths=[80, 120, 130, 90, 75])
    damages_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(damages_table)

    doc.build(story)
    return response



@login_required
def panel_shared_access(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    companies = Company.objects.filter(owner=request.user).order_by('name')
    members = CompanyMember.objects.filter(company__owner=request.user, is_active=True).select_related('company', 'user').order_by('-created_at')
    invites = TaxiShareInvite.objects.filter(invited_by=request.user, is_active=True).select_related('company').order_by('-created_at')

    if request.method == 'POST':
        company_id = request.POST.get('company_id')
        role = request.POST.get('role', CompanyMember.ROLE_VIEWER)

        allowed_roles = [
            CompanyMember.ROLE_PARTNER,
            CompanyMember.ROLE_MANAGER,
            CompanyMember.ROLE_VIEWER,
            CompanyMember.ROLE_ACCOUNTANT,
        ]

        if role not in allowed_roles:
            role = CompanyMember.ROLE_VIEWER

        max_uses = request.POST.get('max_uses', '1')

        try:
            max_uses = int(max_uses)
            if max_uses < 1:
                max_uses = 1
        except ValueError:
            max_uses = 1

        company = get_object_or_404(Company, id=company_id, owner=request.user)

        invite = TaxiShareInvite.objects.create(
            company=company,
            invited_by=request.user,
            role=role,
            max_uses=max_uses,
        )

        audit_event(
            request,
            AuditLog.ACTION_CREATE,
            'sharing',
            'TaxiShareInvite',
            invite.id,
            invite.company.name,
            f'Invitación creada para empresa {invite.company.name} con rol {invite.role}'
        )

        create_notification(
            request.user,
            'Invitación compartida creada',
            f'Se creó un enlace de acceso para {invite.company.name} con rol {invite.get_role_display()}.',
            'success',
            '/panel/shared-access/'
        )

        messages.success(request, 'Enlace compartido creado correctamente.')
        return redirect('/panel/shared-access/')

    return render_panel(request, 'dashboard/shared_access.html', {
        'companies': companies,
        'members': members,
        'invites': invites,
        'roles': [(v, l) for v, l in CompanyMember.ROLE_CHOICES if v != CompanyMember.ROLE_OWNER],
    })


@login_required
def accept_taxi_invite(request, token):
    invite = get_object_or_404(TaxiShareInvite, token=token)

    if not invite.is_valid:
        messages.error(request, 'Esta invitación ya no es válida o ha expirado.')
        return redirect('/panel/')

    if invite.company.owner == request.user:
        messages.info(request, 'Ya eres propietario de esta empresa.')
        return redirect('/panel/')

    permissions = shared_permissions_for_role(invite.role)

    member, created = CompanyMember.objects.get_or_create(
        company=invite.company,
        user=request.user,
        defaults={
            'role': invite.role,
            'is_active': True,
            **permissions,
        }
    )

    if not created:
        member.role = invite.role
        member.is_active = True
        for key, value in permissions.items():
            setattr(member, key, value)
        member.save()

    invite.used_count += 1
    if invite.used_count >= invite.max_uses:
        invite.is_active = False
    invite.save()

    audit_event(
        request,
        AuditLog.ACTION_UPDATE,
        'sharing',
        'CompanyMember',
        invite.company.id,
        invite.company.name,
        f'Usuario aceptó invitación compartida para {invite.company.name} con rol {invite.role}'
    )

    create_notification(
        invite.company.owner,
        'Nuevo socio conectado',
        f'{request.user} aceptó la invitación para acceder a {invite.company.name} como {invite.get_role_display()}.',
        'success',
        '/panel/shared-access/'
    )

    create_notification(
        request.user,
        'Acceso compartido activado',
        f'Ya puedes ver la empresa {invite.company.name} desde tu panel.',
        'success',
        '/panel/'
    )

    messages.success(request, 'Acceso compartido activado correctamente.')
    return redirect('/panel/')


@login_required
def deactivate_taxi_invite(request, invite_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    if request.method != 'POST':
        return redirect('/panel/shared-access/')

    invite = get_object_or_404(
        TaxiShareInvite,
        id=invite_id,
        company__owner=request.user
    )

    invite.is_active = False
    invite.save(update_fields=['is_active'])

    audit_event(
        request,
        AuditLog.ACTION_UPDATE,
        'sharing',
        'TaxiShareInvite',
        invite.id,
        invite.company.name,
        f'Invitación desactivada para empresa {invite.company.name}'
    )

    messages.success(request, 'Invitación desactivada correctamente.')
    return redirect('/panel/shared-access/')


@login_required
def remove_company_member(request, member_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    if request.method != 'POST':
        return redirect('/panel/shared-access/')

    member = get_object_or_404(
        CompanyMember,
        id=member_id,
        company__owner=request.user
    )

    member.is_active = False
    member.save(update_fields=['is_active'])

    audit_event(
        request,
        AuditLog.ACTION_UPDATE,
        'sharing',
        'CompanyMember',
        member.id,
        member.company.name,
        f'Acceso retirado al usuario {member.user} en empresa {member.company.name}'
    )

    create_notification(
        request.user,
        'Acceso retirado',
        f'Se retiró el acceso de {member.user} a {member.company.name}.',
        'warning',
        '/panel/shared-access/'
    )

    create_notification(
        member.user,
        'Acceso desactivado',
        f'Tu acceso a {member.company.name} fue desactivado.',
        'warning',
        '/panel/'
    )

    messages.success(request, 'Acceso del socio desactivado correctamente.')
    return redirect('/panel/shared-access/')


@login_required
def panel_notifications(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    filter_type = request.GET.get('filter', 'all')

    notifications = Notification.objects.filter(user=request.user)

    total_count = notifications.count()
    unread_count = notifications.filter(is_read=False).count()
    read_count = notifications.filter(is_read=True).count()

    if filter_type == 'unread':
        notifications = notifications.filter(is_read=False)
    elif filter_type == 'read':
        notifications = notifications.filter(is_read=True)

    notifications = notifications.order_by('-created_at')

    paginator = Paginator(notifications, 6)
    page_number = request.GET.get('page')
    notifications_page = paginator.get_page(page_number)

    return render_panel(request, 'dashboard/notifications.html', {
        'notifications': notifications_page,
        'total_count': total_count,
        'unread_count': unread_count,
        'read_count': read_count,
        'active_filter': filter_type,
    })


@login_required
def mark_notification_read(request, notification_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    notification = get_object_or_404(Notification, id=notification_id, user=request.user)

    notification.is_read = True
    notification.save(update_fields=['is_read'])

    if notification.link:
        return redirect(notification.link)

    return redirect('/panel/notifications/')


@login_required
def mark_all_notifications_read(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, 'Todas las notificaciones fueron marcadas como leídas.')

    return redirect('/panel/notifications/')


@login_required
def panel_calendar(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    subscription_response = subscription_guard(request)
    if subscription_response:
        return subscription_response

    today = date.today()

    payments = DriverPayment.objects.filter(
        user_company_access_q(request.user)
    ).select_related('company', 'driver', 'vehicle').order_by('payment_date')

    damages = VehicleDamage.objects.filter(
        user_company_access_q(request.user)
    ).select_related('company', 'driver', 'vehicle').order_by('damage_date')

    pending_payments = payments.filter(status__in=['pending', 'partial', 'late'])
    pending_damages = damages.filter(status__in=['pending', 'in_repair'])

    total_pending_debt = sum([p.debt_amount for p in pending_payments])
    total_active_damage_cost = sum([d.estimated_cost for d in pending_damages])
    total_operational_risk = total_pending_debt + total_active_damage_cost

    events = []

    for payment in pending_payments[:80]:
        is_late = payment.payment_date < today or payment.status == 'late'
        events.append({
            'type': 'payment',
            'title': f'Pago pendiente - {payment.driver}',
            'date': payment.payment_date,
            'status': 'Atrasado' if is_late else payment.get_status_display(),
            'amount': payment.debt_amount,
            'vehicle': payment.vehicle,
            'link': '/panel/payments/',
            'level': 'danger' if is_late else 'warning',
        })

    for damage in pending_damages[:80]:
        events.append({
            'type': 'damage',
            'title': f'Daño - {damage.title}',
            'date': damage.damage_date,
            'status': damage.get_status_display(),
            'amount': damage.estimated_cost,
            'vehicle': damage.vehicle,
            'link': '/panel/damages/',
            'level': 'danger' if damage.status == 'pending' else 'warning',
        })

    filter_type = request.GET.get('filter', 'all')

    if filter_type == 'payments':
        events = [e for e in events if e['type'] == 'payment']
    elif filter_type == 'damages':
        events = [e for e in events if e['type'] == 'damage']
    elif filter_type == 'late':
        events = [e for e in events if e['level'] == 'danger']

    events = sorted(events, key=lambda x: x['date'])
    visible_events = events[:8]

    return render_panel(request, 'dashboard/calendar.html', {
        'today': today,
        'events': visible_events,
        'total_events_count': len(events),
        'active_filter': filter_type,
        'pending_payments_count': pending_payments.count(),
        'pending_damages_count': pending_damages.count(),
        'late_payments_count': pending_payments.filter(payment_date__lt=today).count(),
        'total_pending_debt': total_pending_debt,
        'total_active_damage_cost': total_active_damage_cost,
        'total_operational_risk': total_operational_risk,
    })


@login_required
def panel_subscription_renew(request):
    subscription = None

    try:
        subscription = request.user.saas_subscription
    except Exception:
        subscription = None

    payment_settings = None
    try:
        from accounts.models import SaaSPaymentSettings
        payment_settings = SaaSPaymentSettings.get_active()
    except Exception:
        payment_settings = None

    if request.method == 'POST':
        if not subscription:
            messages.error(request, 'No se encontró una suscripción asociada a tu cuenta.')
            return redirect('/panel/subscription/renew/')

        reference = request.POST.get('payment_reference', '').strip()
        receipt = request.FILES.get('payment_receipt')

        if not reference:
            messages.error(request, 'Debes escribir la referencia del pago.')
            return redirect('/panel/subscription/renew/')

        if not receipt:
            messages.error(request, 'Debes subir el comprobante del pago.')
            return redirect('/panel/subscription/renew/')

        subscription.last_payment_reference = reference
        subscription.last_payment_receipt = receipt
        subscription.status = SaaSSubscription.STATUS_PENDING
        subscription.notes = 'Comprobante de renovación enviado por el usuario. Pendiente de revisión administrativa.'
        subscription.save()

        messages.success(
            request,
            'Comprobante enviado correctamente. Tu renovación queda pendiente de revisión administrativa.'
        )
        return redirect('/panel/subscription/renew/')

    return render(request, 'dashboard/subscription_renew.html', {
        'subscription': subscription,
        'payment_settings': payment_settings,
    })

