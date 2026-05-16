from sharing.models import CompanyMember, TaxiShareInvite
from django.contrib import messages
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import csv
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse

from companies.models import Company
from vehicles.models import Vehicle
from drivers.models import Driver
from payments.models import DriverPayment
from damages.models import VehicleDamage
from .forms import VehicleForm, DriverForm, DriverPaymentForm, VehicleDamageForm
from django.db.models import Q, Sum


def user_can_access_panel(user):
    return bool(
        user.is_authenticated
        and user.is_active
        and user.role == 'admin'
        and user.admin_approved
    )


def panel_guard(request):
    user = request.user

    if user.is_superuser and user.is_platform_owner:
        return redirect('/admin/')

    if not user_can_access_panel(user):
        return HttpResponseForbidden(
            'Acceso denegado. Tu cuenta aún no está aprobada como dueño de taxi.'
        )

    return None


def get_user_company(user):
    return Company.objects.filter(owner=user, status='active').first()


def calculate_payment_status(expected_amount, paid_amount):
    if paid_amount <= 0:
        return 'pending'

    if paid_amount >= expected_amount:
        return 'paid'

    return 'partial'



def user_company_access_q(user):
    return (
        Q(company__owner=user) |
        Q(company__members__user=user, company__members__is_active=True)
    )


def user_company_direct_access_q(user):
    return (
        Q(owner=user) |
        Q(members__user=user, members__is_active=True)
    )

@login_required
def panel_home(request):
    from django.utils import timezone

    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    user = request.user

    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)
    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)

    today = timezone.localdate()
    month_start = today.replace(day=1)
    last_7_start = today - timezone.timedelta(days=7)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid

    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost

    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    month_payments = payments.filter(payment_date__gte=month_start, payment_date__lte=today)
    today_payments = payments.filter(payment_date=today)
    last_7_payments = payments.filter(payment_date__gte=last_7_start, payment_date__lte=today)
    month_damages = damages.filter(damage_date__gte=month_start, damage_date__lte=today)

    month_expected = month_payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    month_paid = month_payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    month_debt = month_expected - month_paid
    month_damages_cost = month_damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    month_net_profit = month_paid - month_damages_cost

    today_paid = today_payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    last_7_paid = last_7_payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')

    active_vehicles = vehicles.filter(status='active').count()
    active_drivers = drivers.filter(status='active').count()
    damages_pending = damages.filter(status='pending').count()

    filter_type = request.GET.get('filter', 'month')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    context = {
        'companies_count': companies.count(),
        'vehicles_count': vehicles.count(),
        'drivers_count': drivers.count(),
        'payments_count': payments.count(),
        'damages_count': damages.count(),

        'active_vehicles': active_vehicles,
        'active_drivers': active_drivers,
        'damages_pending': damages_pending,

        'total_expected': total_expected,
        'total_paid': total_paid,
        'total_debt': total_debt,
        'damages_cost': damages_cost,
        'net_profit': net_profit,
        'collection_rate': collection_rate,

        'month_label': today.strftime('%m/%Y'),
        'month_expected': month_expected,
        'month_paid': month_paid,
        'month_debt': month_debt,
        'month_damages_cost': month_damages_cost,
        'month_net_profit': month_net_profit,
        'today_paid': today_paid,
        'last_7_paid': last_7_paid,

        'recent_payments': payments.order_by('-payment_date', '-created_at')[:5],
        'recent_damages': damages.order_by('-damage_date', '-created_at')[:5],

        'active_filter': filter_type,
        'start_date': start_date or '',
        'end_date': end_date or '',

        'chart_paid': float(total_paid),
        'chart_debt': float(total_debt),
        'chart_damages': float(damages_cost),
        'chart_net': float(net_profit),
        'chart_active_vehicles': active_vehicles,
        'chart_active_drivers': active_drivers,
        'chart_damages_pending': damages_pending,
        'chart_payments_paid': payments.filter(status='paid').count(),
        'chart_payments_partial': payments.filter(status='partial').count(),
        'chart_payments_pending': payments.filter(status='pending').count(),
    }

    return render(request, 'dashboard/panel_home.html', context)



@login_required
def panel_vehicles(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        vehicles = vehicles.filter(
            Q(plate_number__icontains=q) |
            Q(brand__icontains=q) |
            Q(model__icontains=q) |
            Q(color__icontains=q)
        )

    if status:
        vehicles = vehicles.filter(status=status).select_related('company')

    return render(request, 'dashboard/panel_vehicles.html', {
        'vehicles': vehicles,
    })


@login_required
def panel_vehicle_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)

        if form.is_valid():
            vehicle = form.save(commit=False)
            vehicle.company = company
            vehicle.created_by = request.user
            vehicle.save()
            return redirect('/panel/vehicles/')
    else:
        form = VehicleForm()

    return render(request, 'dashboard/panel_vehicle_form.html', {
        'form': form,
        'title': 'Registrar nuevo taxi',
        'button_text': 'Guardar taxi',
    })


@login_required
def panel_vehicle_edit(request, vehicle_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    vehicle = get_object_or_404(
        Vehicle.objects.filter(user_company_access_q(request.user)),
        id=vehicle_id
    )

    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES, instance=vehicle)

        if form.is_valid():
            form.save()
            return redirect('/panel/vehicles/')
    else:
        form = VehicleForm(instance=vehicle)

    return render(request, 'dashboard/panel_vehicle_form.html', {
        'form': form,
        'title': 'Editar taxi',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_drivers(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    drivers = Driver.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        drivers = drivers.filter(
            Q(full_name__icontains=q) |
            Q(phone__icontains=q) |
            Q(identity_number__icontains=q) |
            Q(license_number__icontains=q) |
            Q(assigned_vehicle__plate_number__icontains=q)
        )

    if status:
        drivers = drivers.filter(status=status).select_related('company', 'assigned_vehicle')

    return render(request, 'dashboard/panel_drivers.html', {
        'drivers': drivers,
    })


@login_required
def panel_driver_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        form = DriverForm(request.POST, request.FILES, company=company)

        if form.is_valid():
            driver = form.save(commit=False)
            driver.company = company
            driver.created_by = request.user
            driver.save()
            return redirect('/panel/drivers/')
    else:
        form = DriverForm(company=company)

    return render(request, 'dashboard/panel_driver_form.html', {
        'form': form,
        'title': 'Registrar nuevo conductor',
        'button_text': 'Guardar conductor',
    })


@login_required
def panel_driver_edit(request, driver_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    company = get_user_company(request.user)

    driver = get_object_or_404(
        Driver.objects.filter(user_company_access_q(request.user)),
        id=driver_id
    )

    if request.method == 'POST':
        form = DriverForm(request.POST, request.FILES, instance=driver, company=company)

        if form.is_valid():
            form.save()
            return redirect('/panel/drivers/')
    else:
        form = DriverForm(instance=driver, company=company)

    return render(request, 'dashboard/panel_driver_form.html', {
        'form': form,
        'title': 'Editar conductor',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_payments(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    payments = DriverPayment.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        payments = payments.filter(
            Q(driver__full_name__icontains=q) |
            Q(vehicle__plate_number__icontains=q) |
            Q(notes__icontains=q)
        )

    if status:
        payments = payments.filter(status=status).select_related(
        'company',
        'driver',
        'vehicle',
    ).order_by('-payment_date', '-created_at')

    return render(request, 'dashboard/panel_payments.html', {
        'payments': payments,
    })


@login_required
def panel_payment_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        form = DriverPaymentForm(request.POST, company=company)

        if form.is_valid():
            payment = form.save(commit=False)
            payment.company = company
            payment.registered_by = request.user
            payment.status = calculate_payment_status(
                payment.expected_amount,
                payment.paid_amount
            )
            payment.save()
            return redirect('/panel/payments/')
    else:
        form = DriverPaymentForm(company=company)

    return render(request, 'dashboard/panel_payment_form.html', {
        'form': form,
        'title': 'Registrar pago',
        'button_text': 'Guardar pago',
    })

@login_required
def panel_payment_edit(request, payment_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    company = get_user_company(request.user)

    payment = get_object_or_404(
        DriverPayment.objects.filter(user_company_access_q(request.user)),
        id=payment_id
    )

    if request.method == 'POST':
        form = DriverPaymentForm(request.POST, instance=payment, company=company)

        if form.is_valid():
            payment = form.save(commit=False)
            payment.company = company
            payment.registered_by = request.user
            payment.status = calculate_payment_status(
                payment.expected_amount,
                payment.paid_amount
            )
            payment.save()
            return redirect('/panel/payments/')
    else:
        form = DriverPaymentForm(instance=payment, company=company)

    return render(request, 'dashboard/panel_payment_form.html', {
        'form': form,
        'title': 'Editar pago',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_damages(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user))

    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if q:
        damages = damages.filter(
            Q(title__icontains=q) |
            Q(description__icontains=q) |
            Q(vehicle__plate_number__icontains=q) |
            Q(driver__full_name__icontains=q)
        )

    if status:
        damages = damages.filter(status=status).select_related(
        'company',
        'driver',
        'vehicle',
    ).order_by('-damage_date', '-created_at')

    return render(request, 'dashboard/panel_damages.html', {
        'damages': damages,
    })


@login_required
def panel_damage_create(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    company = get_user_company(request.user)

    if not company:
        return HttpResponseForbidden(
            'No tienes una empresa activa asignada. Contacta con el soporte de TaxiGE.'
        )

    if request.method == 'POST':
        form = VehicleDamageForm(request.POST, request.FILES, company=company)

        if form.is_valid():
            damage = form.save(commit=False)
            damage.company = company
            damage.registered_by = request.user
            damage.save()
            return redirect('/panel/damages/')
    else:
        form = VehicleDamageForm(company=company)

    return render(request, 'dashboard/panel_damage_form.html', {
        'form': form,
        'title': 'Registrar daño',
        'button_text': 'Guardar daño',
    })

@login_required
def panel_damage_edit(request, damage_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    company = get_user_company(request.user)

    damage = get_object_or_404(
        VehicleDamage.objects.filter(user_company_access_q(request.user)),
        id=damage_id
    )

    if request.method == 'POST':
        form = VehicleDamageForm(request.POST, request.FILES, instance=damage, company=company)

        if form.is_valid():
            damage = form.save(commit=False)
            damage.company = company
            damage.registered_by = request.user
            damage.save()
            return redirect('/panel/damages/')
    else:
        form = VehicleDamageForm(instance=damage, company=company)

    return render(request, 'dashboard/panel_damage_form.html', {
        'form': form,
        'title': 'Editar daño',
        'button_text': 'Guardar cambios',
    })


@login_required
def panel_vehicle_delete(request, vehicle_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    vehicle = get_object_or_404(Vehicle.objects.filter(user_company_access_q(request.user)), id=vehicle_id)

    if request.method == 'POST':
        vehicle.delete()
        return redirect('/panel/vehicles/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar taxi',
        'object_name': vehicle.plate_number,
        'cancel_url': '/panel/vehicles/',
    })


@login_required
def panel_driver_delete(request, driver_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    driver = get_object_or_404(Driver.objects.filter(user_company_access_q(request.user)), id=driver_id)

    if request.method == 'POST':
        driver.delete()
        return redirect('/panel/drivers/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar conductor',
        'object_name': driver.full_name,
        'cancel_url': '/panel/drivers/',
    })


@login_required
def panel_payment_delete(request, payment_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    payment = get_object_or_404(DriverPayment.objects.filter(user_company_access_q(request.user)), id=payment_id)

    if request.method == 'POST':
        payment.delete()
        return redirect('/panel/payments/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar pago',
        'object_name': f'{payment.driver.full_name} - {payment.payment_date}',
        'cancel_url': '/panel/payments/',
    })


@login_required
def panel_damage_delete(request, damage_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    damage = get_object_or_404(VehicleDamage.objects.filter(user_company_access_q(request.user)), id=damage_id)

    if request.method == 'POST':
        damage.delete()
        return redirect('/panel/damages/')

    return render(request, 'dashboard/panel_confirm_delete.html', {
        'title': 'Eliminar daño',
        'object_name': damage.title,
        'cancel_url': '/panel/damages/',
    })


@login_required
def api_panel_summary(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    user = request.user

    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)
    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or 0
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or 0
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost

    if total_expected > 0:
        collection_rate = round((total_paid / total_expected) * 100, 2)
    else:
        collection_rate = Decimal('0')
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or 0

    data = {
        'companies_count': companies.count(),
        'vehicles_count': vehicles.count(),
        'active_vehicles': vehicles.filter(status='active').count(),
        'drivers_count': drivers.count(),
        'active_drivers': drivers.filter(status='active').count(),
        'payments_count': payments.count(),
        'damages_count': damages.count(),
        'damages_pending': damages.filter(status='pending').count(),
        'total_expected': str(total_expected),
        'total_paid': str(total_paid),
        'total_debt': str(total_debt),
        'damages_cost': str(damages_cost),
    }

    return JsonResponse({
        'ok': True,
        'data': data,
    })


@login_required
def api_panel_vehicles(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user)).select_related('company').order_by('-created_at')

    data = []
    for v in vehicles:
        data.append({
            'id': v.id,
            'plate_number': v.plate_number,
            'brand': v.brand,
            'model': v.model,
            'color': v.color,
            'year': v.year,
            'daily_target_amount': str(v.daily_target_amount),
            'status': v.status,
            'status_display': v.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


@login_required
def api_panel_drivers(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    drivers = Driver.objects.filter(user_company_access_q(request.user)).select_related('assigned_vehicle').order_by('-created_at')

    data = []
    for d in drivers:
        data.append({
            'id': d.id,
            'full_name': d.full_name,
            'phone': d.phone,
            'assigned_vehicle': d.assigned_vehicle.plate_number if d.assigned_vehicle else None,
            'daily_payment_amount': str(d.daily_payment_amount),
            'payment_day': d.payment_day,
            'status': d.status,
            'status_display': d.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


@login_required
def api_panel_payments(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    payments = DriverPayment.objects.filter(user_company_access_q(request.user)).select_related('driver', 'vehicle').order_by('-payment_date', '-created_at')[:100]

    data = []
    for p in payments:
        data.append({
            'id': p.id,
            'driver': p.driver.full_name,
            'vehicle': p.vehicle.plate_number if p.vehicle else None,
            'payment_date': str(p.payment_date),
            'expected_amount': str(p.expected_amount),
            'paid_amount': str(p.paid_amount),
            'debt_amount': str(p.debt_amount),
            'status': p.status,
            'status_display': p.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


@login_required
def api_panel_damages(request):
    guard_response = panel_guard(request)
    if guard_response:
        return JsonResponse({'ok': False, 'error': 'Acceso denegado'}, status=403)

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user)).select_related('driver', 'vehicle').order_by('-damage_date', '-created_at')[:100]

    data = []
    for d in damages:
        data.append({
            'id': d.id,
            'vehicle': d.vehicle.plate_number if d.vehicle else None,
            'driver': d.driver.full_name if d.driver else None,
            'title': d.title,
            'damage_date': str(d.damage_date),
            'estimated_cost': str(d.estimated_cost),
            'final_cost': str(d.final_cost),
            'status': d.status,
            'status_display': d.get_status_display(),
        })

    return JsonResponse({'ok': True, 'data': data})


def make_csv_response(filename, headers, rows):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(headers)

    for row in rows:
        writer.writerow(row)

    return response


@login_required
def export_vehicles_csv(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user)).select_related('company').order_by('-created_at')

    rows = []
    for v in vehicles:
        rows.append([
            v.id,
            v.company.name,
            v.plate_number,
            v.internal_code,
            v.brand,
            v.model,
            v.color,
            v.year,
            v.daily_target_amount,
            v.get_status_display(),
            v.created_at,
        ])

    return make_csv_response(
        'taxige_taxis.csv',
        ['ID', 'Empresa', 'Matrícula', 'Código interno', 'Marca', 'Modelo', 'Color', 'Año', 'Objetivo diario', 'Estado', 'Creado'],
        rows
    )


@login_required
def export_drivers_csv(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    drivers = Driver.objects.filter(user_company_access_q(request.user)).select_related('company', 'assigned_vehicle').order_by('-created_at')

    rows = []
    for d in drivers:
        rows.append([
            d.id,
            d.company.name,
            d.full_name,
            d.phone,
            d.identity_number,
            d.license_number,
            d.assigned_vehicle.plate_number if d.assigned_vehicle else '',
            d.daily_payment_amount,
            d.payment_day,
            d.get_status_display(),
            d.created_at,
        ])

    return make_csv_response(
        'taxige_conductores.csv',
        ['ID', 'Empresa', 'Conductor', 'Teléfono', 'Documento', 'Licencia', 'Taxi asignado', 'Pago diario', 'Día de pago', 'Estado', 'Creado'],
        rows
    )


@login_required
def export_payments_csv(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    payments = DriverPayment.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-payment_date', '-created_at')

    rows = []
    for p in payments:
        rows.append([
            p.id,
            p.company.name,
            p.driver.full_name,
            p.vehicle.plate_number if p.vehicle else '',
            p.payment_date,
            p.expected_amount,
            p.paid_amount,
            p.debt_amount,
            p.get_status_display(),
            p.notes,
            p.created_at,
        ])

    return make_csv_response(
        'taxige_pagos.csv',
        ['ID', 'Empresa', 'Conductor', 'Taxi', 'Fecha', 'Esperado', 'Pagado', 'Deuda', 'Estado', 'Notas', 'Creado'],
        rows
    )


@login_required
def export_damages_csv(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-damage_date', '-created_at')

    rows = []
    for d in damages:
        rows.append([
            d.id,
            d.company.name,
            d.vehicle.plate_number if d.vehicle else '',
            d.driver.full_name if d.driver else '',
            d.title,
            d.damage_date,
            d.estimated_cost,
            d.final_cost,
            d.get_status_display(),
            d.description,
            d.created_at,
        ])

    return make_csv_response(
        'taxige_danos.csv',
        ['ID', 'Empresa', 'Taxi', 'Conductor', 'Daño', 'Fecha', 'Costo estimado', 'Costo final', 'Estado', 'Descripción', 'Creado'],
        rows
    )



def make_pdf_response(filename, title, headers, rows):
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak
    from django.utils import timezone

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    generated_at = timezone.localtime().strftime('%d/%m/%Y %H:%M')

    header_data = [
        [
            Paragraph('<b>TaxiGE Platform</b><br/><font size="8">Sistema profesional de gestión de taxis</font>', styles['Normal']),
            Paragraph(f'<b>{title}</b><br/><font size="8">Generado: {generated_at}</font>', styles['Normal']),
        ]
    ]

    header_table = Table(header_data, colWidths=[130 * mm, 130 * mm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#111827')),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#334155')),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 14))

    summary_data = [
        ['Total de registros', str(len(rows)), 'Formato', 'PDF corporativo'],
    ]

    summary_table = Table(summary_data, colWidths=[45 * mm, 35 * mm, 35 * mm, 55 * mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0F172A')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 14))

    safe_rows = rows if rows else [['Sin datos'] + ['' for _ in headers[1:]]]
    data = [headers] + safe_rows

    available_width = landscape(A4)[0] - (32 * mm)
    col_count = len(headers)
    col_widths = [available_width / col_count for _ in headers]

    table = Table(data, repeatRows=1, colWidths=col_widths)

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(
        '<font size="8" color="#64748B">Documento generado automáticamente por TaxiGE Platform. Uso interno y administrativo.</font>',
        styles['Normal']
    ))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#64748B'))
        canvas.drawString(16 * mm, 10 * mm, 'TaxiGE Platform')
        canvas.drawRightString(281 * mm, 10 * mm, f'Página {doc.page}')
        canvas.restoreState()

    doc.build(elements, onFirstPage=footer, onLaterPages=footer)

    return response


@login_required
def export_vehicles_pdf(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    vehicles = Vehicle.objects.filter(user_company_access_q(request.user)).select_related('company').order_by('-created_at')

    rows = [[
        str(v.id),
        v.company.name,
        v.plate_number,
        v.brand,
        v.model,
        v.color,
        str(v.daily_target_amount),
        v.get_status_display(),
    ] for v in vehicles]

    return make_pdf_response(
        'taxige_taxis.pdf',
        'Reporte de taxis',
        ['ID', 'Empresa', 'Matrícula', 'Marca', 'Modelo', 'Color', 'Objetivo diario', 'Estado'],
        rows
    )


@login_required
def export_drivers_pdf(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    drivers = Driver.objects.filter(user_company_access_q(request.user)).select_related('company', 'assigned_vehicle').order_by('-created_at')

    rows = [[
        str(d.id),
        d.company.name,
        d.full_name,
        d.phone,
        d.assigned_vehicle.plate_number if d.assigned_vehicle else '',
        str(d.daily_payment_amount),
        d.payment_day,
        d.get_status_display(),
    ] for d in drivers]

    return make_pdf_response(
        'taxige_conductores.pdf',
        'Reporte de conductores',
        ['ID', 'Empresa', 'Conductor', 'Teléfono', 'Taxi', 'Pago diario', 'Día pago', 'Estado'],
        rows
    )


@login_required
def export_payments_pdf(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    payments = DriverPayment.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-payment_date', '-created_at')

    rows = [[
        str(p.id),
        p.driver.full_name,
        p.vehicle.plate_number if p.vehicle else '',
        str(p.payment_date),
        str(p.expected_amount),
        str(p.paid_amount),
        str(p.debt_amount),
        p.get_status_display(),
    ] for p in payments]

    return make_pdf_response(
        'taxige_pagos.pdf',
        'Reporte de pagos',
        ['ID', 'Conductor', 'Taxi', 'Fecha', 'Esperado', 'Pagado', 'Deuda', 'Estado'],
        rows
    )


@login_required
def export_damages_pdf(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    damages = VehicleDamage.objects.filter(user_company_access_q(request.user)).select_related('company', 'driver', 'vehicle').order_by('-damage_date', '-created_at')

    rows = [[
        str(d.id),
        d.vehicle.plate_number if d.vehicle else '',
        d.driver.full_name if d.driver else '',
        d.title,
        str(d.damage_date),
        str(d.estimated_cost),
        str(d.final_cost),
        d.get_status_display(),
    ] for d in damages]

    return make_pdf_response(
        'taxige_danos.pdf',
        'Reporte de daños',
        ['ID', 'Taxi', 'Conductor', 'Daño', 'Fecha', 'Estimado', 'Final', 'Estado'],
        rows
    )


@login_required
def dashboard_export_excel(request):
    from django.http import HttpResponse
    from django.db.models import Sum
    from decimal import Decimal
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from payments.models import DriverPayment
    from damages.models import VehicleDamage
    from companies.models import Company
    from vehicles.models import Vehicle
    from drivers.models import Driver

    user = request.user

    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)
    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost
    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    wb = Workbook()

    dark = "0F172A"
    yellow = "FACC15"
    green = "22C55E"
    gray = "E2E8F0"

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ======================
    # HOJA RESUMEN
    # ======================
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "Reporte Ejecutivo TaxiGE"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Usuario: {user.username}"
    ws["A3"] = "Sistema: TaxiGE Platform"

    resumen = [
        ["Indicador", "Valor"],
        ["Ingresos cobrados", float(total_paid)],
        ["Deuda pendiente", float(total_debt)],
        ["Costo daños", float(damages_cost)],
        ["Beneficio neto", float(net_profit)],
        ["Ratio de cobro", f"{collection_rate}%"],
        ["Empresas", companies.count()],
        ["Taxis activos", vehicles.filter(status='active').count()],
        ["Conductores activos", drivers.filter(status='active').count()],
        ["Daños pendientes", damages.filter(status='pending').count()],
    ]

    start_row = 5
    for r, row in enumerate(resumen, start=start_row):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r == start_row:
                cell.font = Font(bold=True, color="111827")
                cell.fill = PatternFill("solid", fgColor=yellow)

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 26

    # ======================
    # HOJA PAGOS
    # ======================
    ws2 = wb.create_sheet("Pagos")
    headers = ["Conductor", "Taxi", "Fecha", "Esperado", "Pagado", "Deuda", "Estado"]

    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r, pay in enumerate(payments.order_by("-payment_date", "-created_at"), 2):
        row = [
            getattr(pay.driver, "full_name", ""),
            getattr(pay.vehicle, "plate_number", ""),
            str(pay.payment_date),
            float(pay.expected_amount),
            float(pay.paid_amount),
            float(pay.debt_amount),
            pay.get_status_display(),
        ]
        for c, value in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ======================
    # HOJA DAÑOS
    # ======================
    ws3 = wb.create_sheet("Daños")
    headers = ["Taxi", "Conductor", "Daño", "Fecha", "Costo estimado", "Estado"]

    for c, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r, dmg in enumerate(damages.order_by("-damage_date", "-created_at"), 2):
        row = [
            getattr(dmg.vehicle, "plate_number", ""),
            getattr(dmg.driver, "full_name", ""),
            dmg.title,
            str(dmg.damage_date),
            float(dmg.estimated_cost),
            dmg.get_status_display(),
        ]
        for c, value in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 24

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="taxige_reporte_ejecutivo.xlsx"'
    wb.save(response)
    return response


@login_required
def dashboard_export_pdf(request):
    from django.http import HttpResponse
    from django.db.models import Sum
    from decimal import Decimal
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from payments.models import DriverPayment
    from damages.models import VehicleDamage
    from companies.models import Company
    from vehicles.models import Vehicle
    from drivers.models import Driver

    user = request.user

    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)
    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost
    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="taxige_reporte_ejecutivo.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("<b>Reporte Ejecutivo TaxiGE</b>", styles["Title"]))
    story.append(Paragraph(f"Usuario: {user.username}", styles["Normal"]))
    story.append(Spacer(1, 14))

    resumen = [
        ["Indicador", "Valor"],
        ["Ingresos cobrados", f"{total_paid} XAF"],
        ["Deuda pendiente", f"{total_debt} XAF"],
        ["Costo daños", f"{damages_cost} XAF"],
        ["Beneficio neto", f"{net_profit} XAF"],
        ["Ratio de cobro", f"{collection_rate}%"],
        ["Empresas", companies.count()],
        ["Taxis activos", vehicles.filter(status='active').count()],
        ["Conductores activos", drivers.filter(status='active').count()],
        ["Daños pendientes", damages.filter(status='pending').count()],
    ]

    table = Table(resumen, colWidths=[220, 220])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Últimos pagos</b>", styles["Heading2"]))
    pagos_data = [["Conductor", "Taxi", "Fecha", "Pagado", "Estado"]]
    for pay in payments.order_by("-payment_date", "-created_at")[:8]:
        pagos_data.append([
            getattr(pay.driver, "full_name", ""),
            getattr(pay.vehicle, "plate_number", ""),
            str(pay.payment_date),
            f"{pay.paid_amount} XAF",
            pay.get_status_display(),
        ])

    pagos_table = Table(pagos_data, colWidths=[120, 80, 85, 95, 95])
    pagos_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#facc15")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(pagos_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Últimos daños</b>", styles["Heading2"]))
    damages_data = [["Taxi", "Conductor", "Daño", "Costo", "Estado"]]
    for dmg in damages.order_by("-damage_date", "-created_at")[:8]:
        damages_data.append([
            getattr(dmg.vehicle, "plate_number", ""),
            getattr(dmg.driver, "full_name", ""),
            dmg.title,
            f"{dmg.estimated_cost} XAF",
            dmg.get_status_display(),
        ])

    damages_table = Table(damages_data, colWidths=[80, 120, 130, 90, 75])
    damages_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(damages_table)

    doc.build(story)
    return response



@login_required
def panel_shared_access(request):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    companies = Company.objects.filter(owner=request.user).order_by('name')
    members = CompanyMember.objects.filter(company__owner=request.user).select_related('company', 'user').order_by('-created_at')
    invites = TaxiShareInvite.objects.filter(invited_by=request.user, is_active=True).select_related('company').order_by('-created_at')

    if request.method == 'POST':
        company_id = request.POST.get('company_id')
        role = request.POST.get('role', CompanyMember.ROLE_VIEWER)
        max_uses = request.POST.get('max_uses', '1')

        try:
            max_uses = int(max_uses)
            if max_uses < 1:
                max_uses = 1
        except ValueError:
            max_uses = 1

        company = get_object_or_404(Company, id=company_id, owner=request.user)

        invite = TaxiShareInvite.objects.create(
            company=company,
            invited_by=request.user,
            role=role,
            max_uses=max_uses,
        )

        messages.success(request, 'Enlace compartido creado correctamente.')
        return redirect('/panel/shared-access/')

    return render(request, 'dashboard/shared_access.html', {
        'companies': companies,
        'members': members,
        'invites': invites,
        'roles': [(v, l) for v, l in CompanyMember.ROLE_CHOICES if v != CompanyMember.ROLE_OWNER],
    })


@login_required
def accept_taxi_invite(request, token):
    invite = get_object_or_404(TaxiShareInvite, token=token)

    if not invite.is_valid:
        messages.error(request, 'Esta invitación ya no es válida o ha expirado.')
        return redirect('/panel/')

    if invite.company.owner == request.user:
        messages.info(request, 'Ya eres propietario de esta empresa.')
        return redirect('/panel/')

    CompanyMember.objects.get_or_create(
        company=invite.company,
        user=request.user,
        defaults={
            'role': invite.role,
            'is_active': True,
            'can_view_payments': True,
            'can_view_damages': True,
            'can_export_reports': True,
        }
    )

    invite.used_count += 1
    if invite.used_count >= invite.max_uses:
        invite.is_active = False
    invite.save()

    messages.success(request, 'Acceso compartido activado correctamente.')
    return redirect('/panel/')


@login_required
def deactivate_taxi_invite(request, invite_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    if request.method != 'POST':
        return redirect('/panel/shared-access/')

    invite = get_object_or_404(
        TaxiShareInvite,
        id=invite_id,
        company__owner=request.user
    )

    invite.is_active = False
    invite.save(update_fields=['is_active'])

    messages.success(request, 'Invitación desactivada correctamente.')
    return redirect('/panel/shared-access/')


@login_required
def remove_company_member(request, member_id):
    guard_response = panel_guard(request)
    if guard_response:
        return guard_response

    if request.method != 'POST':
        return redirect('/panel/shared-access/')

    member = get_object_or_404(
        CompanyMember,
        id=member_id,
        company__owner=request.user
    )

    member.is_active = False
    member.save(update_fields=['is_active'])

    messages.success(request, 'Acceso del socio desactivado correctamente.')
    return redirect('/panel/shared-access/')
