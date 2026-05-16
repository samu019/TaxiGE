from pathlib import Path

# =========================
# 1. ACTUALIZAR VIEWS
# =========================
views_path = Path("dashboard/views.py")
views = views_path.read_text(encoding="utf-8-sig").replace("\ufeff", "")

if "def dashboard_export_excel" not in views:
    views += r'''

@login_required
def dashboard_export_excel(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.db.models import Sum
    from decimal import Decimal
    from payments.models import DriverPayment
    from damages.models import VehicleDamage

    user = request.user
    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard TaxiGE"

    ws["A1"] = "Reporte Ejecutivo TaxiGE"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="FACC15")

    data = [
        ["Indicador", "Valor"],
        ["Ingresos cobrados", float(total_paid)],
        ["Deuda pendiente", float(total_debt)],
        ["Costo daños", float(damages_cost)],
        ["Beneficio neto", float(net_profit)],
    ]

    for row in data:
        ws.append(row)

    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="taxige_dashboard.xlsx"'
    wb.save(response)
    return response


@login_required
def dashboard_export_pdf(request):
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from django.db.models import Sum
    from decimal import Decimal
    from payments.models import DriverPayment
    from damages.models import VehicleDamage

    user = request.user
    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="taxige_dashboard.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    pdf.setFont("Helvetica-Bold", 20)
    pdf.drawString(50, height - 60, "Reporte Ejecutivo TaxiGE")

    pdf.setFont("Helvetica", 12)
    y = height - 110

    rows = [
        ("Usuario", user.username),
        ("Ingresos cobrados", f"{total_paid} XAF"),
        ("Deuda pendiente", f"{total_debt} XAF"),
        ("Costo daños", f"{damages_cost} XAF"),
        ("Beneficio neto", f"{net_profit} XAF"),
    ]

    for label, value in rows:
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, y, label + ":")
        pdf.setFont("Helvetica", 12)
        pdf.drawString(210, y, str(value))
        y -= 28

    pdf.showPage()
    pdf.save()
    return response
'''

views_path.write_text(views, encoding="utf-8")


# =========================
# 2. ACTUALIZAR URLS
# =========================
urls_path = Path("dashboard/urls.py")
urls = urls_path.read_text(encoding="utf-8-sig").replace("\ufeff", "")

if "dashboard_export_excel" not in urls:
    urls = urls.replace(
        "urlpatterns = [",
        """urlpatterns = [
    path('export/excel/', views.dashboard_export_excel, name='dashboard_export_excel'),
    path('export/pdf/', views.dashboard_export_pdf, name='dashboard_export_pdf'),"""
    )

urls_path.write_text(urls, encoding="utf-8")


# =========================
# 3. DASHBOARD PREMIUM
# =========================
template_path = Path("templates/dashboard/panel_home.html")

template = r'''{% extends 'dashboard/panel_base.html' %}

{% block title %}Panel ejecutivo - TaxiGE{% endblock %}

{% block content %}

<style>
.dashboard-actions {
    display: flex;
    gap: 10px;
    flex-wrap: wrap;
}

.export-btn {
    min-height: 42px;
    border-radius: 14px;
    padding: 0 16px;
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background: linear-gradient(135deg, #f59e0b, #facc15);
    color: #111827;
    font-weight: 1000;
    border: none;
}

.filter-panel,
.metric-card,
.premium-section {
    background: rgba(15,23,42,.92);
    border: 1px solid rgba(148,163,184,.22);
    box-shadow: 0 20px 50px rgba(0,0,0,.20);
}

.filter-panel {
    border-radius: 24px;
    padding: 18px;
    margin-bottom: 22px;
}

.filter-bar {
    display: grid;
    grid-template-columns: 1.5fr 1fr 1fr auto;
    gap: 12px;
    align-items: center;
}

.filter-bar select,
.filter-bar input {
    min-height: 48px;
    border-radius: 16px;
    border: 1px solid rgba(148,163,184,.28);
    background: rgba(2,6,23,.72);
    color: #f8fafc;
    padding: 0 14px;
    font-weight: 900;
}

.btn-filter {
    min-height: 48px;
    border: none;
    border-radius: 16px;
    padding: 0 20px;
    background: linear-gradient(135deg, #f59e0b, #facc15);
    color: #111827;
    font-weight: 1000;
    cursor: pointer;
}

.metric-grid {
    display: grid;
    grid-template-columns: repeat(5, minmax(0, 1fr));
    gap: 16px;
    margin-bottom: 22px;
}

.metric-card {
    border-radius: 24px;
    padding: 18px;
    position: relative;
    overflow: hidden;
}

.metric-card::before {
    content: "";
    position: absolute;
    inset: 0 0 auto 0;
    height: 5px;
    background: linear-gradient(90deg, #f59e0b, #22c55e);
}

.metric-top {
    display: flex;
    justify-content: space-between;
    gap: 10px;
    align-items: center;
}

.metric-icon {
    width: 38px;
    height: 38px;
    border-radius: 14px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    background: rgba(245,158,11,.14);
    color: #facc15;
    font-weight: 1000;
}

.metric-label {
    color: #cbd5e1;
    font-size: 13px;
    font-weight: 1000;
}

.metric-value {
    margin-top: 12px;
    color: #f8fafc;
    font-size: 30px;
    line-height: 1;
    font-weight: 1000;
}

.metric-note {
    margin-top: 9px;
    color: #94a3b8;
    font-size: 12px;
    font-weight: 800;
}

.trend-up {
    color: #22c55e;
}

.trend-down {
    color: #ef4444;
}

.trend-neutral {
    color: #facc15;
}

.premium-section {
    border-radius: 26px;
    margin-bottom: 20px;
    overflow: hidden;
}

.premium-section summary {
    list-style: none;
    cursor: pointer;
    padding: 20px 24px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    color: #f8fafc;
    font-size: 22px;
    font-weight: 1000;
}

.premium-section summary::-webkit-details-marker {
    display: none;
}

.premium-section summary::after {
    content: "›";
    color: #f59e0b;
    font-size: 34px;
    font-weight: 1000;
    transform: rotate(90deg);
    transition: .22s ease;
}

.premium-section[open] summary::after {
    transform: rotate(-90deg);
    color: #facc15;
}

.section-body {
    padding: 0 24px 24px;
}

.compact-grid {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 14px;
}

.chart-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 18px;
}

.chart-card {
    background: rgba(2,6,23,.45);
    border: 1px solid rgba(148,163,184,.20);
    border-radius: 22px;
    padding: 18px;
    min-height: 310px;
}

.chart-card h3 {
    margin: 0 0 12px;
    color: #f8fafc;
    font-size: 16px;
    font-weight: 1000;
}

.table-wrap {
    overflow-x: auto;
    border-radius: 20px;
    border: 1px solid rgba(148,163,184,.24);
}

.table-wrap table {
    width: 100%;
}

.badge,
.status-paid,
.status-partial,
.status-pending {
    border-radius: 999px;
    padding: 7px 12px;
    font-weight: 1000;
    display: inline-flex;
}

.status-paid {
    background: rgba(34,197,94,.16);
    color: #22c55e;
}

.status-partial {
    background: rgba(245,158,11,.16);
    color: #facc15;
}

.status-pending {
    background: rgba(239,68,68,.16);
    color: #fb7185;
}

.live-alert {
    display: none;
    margin-bottom: 18px;
    border-radius: 20px;
    padding: 14px 18px;
    background: rgba(245,158,11,.14);
    border: 1px solid rgba(245,158,11,.35);
    color: #facc15;
    font-weight: 1000;
}

.live-alert.show {
    display: block;
}

@media (max-width: 1200px) {
    .metric-grid {
        grid-template-columns: repeat(2, minmax(0, 1fr));
    }

    .compact-grid,
    .chart-grid {
        grid-template-columns: 1fr;
    }
}

@media (max-width: 760px) {
    .filter-bar,
    .metric-grid,
    .compact-grid,
    .chart-grid {
        grid-template-columns: 1fr;
    }

    .dashboard-actions {
        width: 100%;
    }

    .export-btn {
        width: 100%;
        justify-content: center;
    }

    .metric-value {
        font-size: 24px;
    }

    .premium-section summary {
        font-size: 18px;
        padding: 18px;
    }
}
</style>

<div class="dashboard-hero">
    <div>
        <h1 class="page-title">Dashboard ejecutivo</h1>
        <p class="page-subtitle">Resumen financiero, mensual y operativo de tu empresa de taxis.</p>
    </div>

    <div class="dashboard-actions">
        <a class="export-btn" href="{% url 'dashboard_export_pdf' %}">PDF</a>
        <a class="export-btn" href="{% url 'dashboard_export_excel' %}">Excel</a>
    </div>
</div>

<div id="liveAlert" class="live-alert">
    Hay indicadores pendientes que requieren revisión.
</div>

<div class="filter-panel">
    <form method="get" class="filter-bar">
        <select name="filter">
            <option value="month" {% if active_filter == 'month' %}selected{% endif %}>Este mes</option>
            <option value="today" {% if active_filter == 'today' %}selected{% endif %}>Hoy</option>
            <option value="week" {% if active_filter == 'week' %}selected{% endif %}>Últimos 7 días</option>
            <option value="custom" {% if active_filter == 'custom' %}selected{% endif %}>Personalizado</option>
        </select>

        <input type="date" name="start_date" value="{{ start_date }}">
        <input type="date" name="end_date" value="{{ end_date }}">

        <button type="submit" class="btn-filter">Aplicar filtro</button>
    </form>
</div>

<div class="metric-grid">
    <div class="metric-card">
        <div class="metric-top">
            <div class="metric-label">Ingresos cobrados</div>
            <div class="metric-icon">XAF</div>
        </div>
        <div class="metric-value counter" data-value="{{ total_paid }}">0</div>
        <div class="metric-note trend-up">↑ Total recibido</div>
    </div>

    <div class="metric-card">
        <div class="metric-top">
            <div class="metric-label">Deuda pendiente</div>
            <div class="metric-icon">D</div>
        </div>
        <div class="metric-value counter" data-value="{{ total_debt }}">0</div>
        <div class="metric-note trend-down">↓ Por cobrar</div>
    </div>

    <div class="metric-card">
        <div class="metric-top">
            <div class="metric-label">Costo daños</div>
            <div class="metric-icon">R</div>
        </div>
        <div class="metric-value counter" data-value="{{ damages_cost }}">0</div>
        <div class="metric-note trend-neutral">Gastos estimados</div>
    </div>

    <div class="metric-card">
        <div class="metric-top">
            <div class="metric-label">Beneficio neto</div>
            <div class="metric-icon">N</div>
        </div>
        <div class="metric-value counter" data-value="{{ net_profit }}">0</div>
        <div class="metric-note trend-up">↑ Ingresos menos daños</div>
    </div>

    <div class="metric-card">
        <div class="metric-top">
            <div class="metric-label">Ratio de cobro</div>
            <div class="metric-icon">%</div>
        </div>
        <div class="metric-value">{{ collection_rate }}%</div>
        <div class="metric-note trend-up">Rendimiento financiero</div>
    </div>
</div>

<details class="premium-section" open>
    <summary>Gráficos ejecutivos</summary>
    <div class="section-body">
        <div class="chart-grid">
            <div class="chart-card">
                <h3>Finanzas principales</h3>
                <canvas id="financeChart"></canvas>
            </div>
            <div class="chart-card">
                <h3>Operación actual</h3>
                <canvas id="operationChart"></canvas>
            </div>
        </div>
    </div>
</details>

<details class="premium-section">
    <summary>Resumen mensual - 05/2026</summary>
    <div class="section-body">
        <div class="compact-grid">
            <div class="metric-card"><div class="metric-label">Ingresos del mes</div><div class="metric-value">{{ total_paid }} XAF</div></div>
            <div class="metric-card"><div class="metric-label">Deuda del mes</div><div class="metric-value">{{ total_debt }} XAF</div></div>
            <div class="metric-card"><div class="metric-label">Daños del mes</div><div class="metric-value">{{ damages_cost }} XAF</div></div>
            <div class="metric-card"><div class="metric-label">Neto mensual</div><div class="metric-value">{{ net_profit }} XAF</div></div>
        </div>
    </div>
</details>

<details class="premium-section">
    <summary>Operación actual</summary>
    <div class="section-body">
        <div class="compact-grid">
            <div class="metric-card"><div class="metric-label">Empresas</div><div class="metric-value">{{ companies_count }}</div></div>
            <div class="metric-card"><div class="metric-label">Taxis activos</div><div class="metric-value">{{ active_vehicles }}</div></div>
            <div class="metric-card"><div class="metric-label">Conductores activos</div><div class="metric-value">{{ active_drivers }}</div></div>
            <div class="metric-card"><div class="metric-label">Daños pendientes</div><div class="metric-value">{{ damages_pending }}</div></div>
        </div>
    </div>
</details>

<details class="premium-section">
    <summary>Últimos pagos</summary>
    <div class="section-body">
        {% if recent_payments %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Conductor</th>
                        <th>Taxi</th>
                        <th>Fecha</th>
                        <th>Esperado</th>
                        <th>Pagado</th>
                        <th>Deuda</th>
                        <th>Estado</th>
                    </tr>
                </thead>
                <tbody>
                    {% for payment in recent_payments %}
                    <tr>
                        <td>{{ payment.driver.full_name }}</td>
                        <td>{{ payment.vehicle.plate_number }}</td>
                        <td>{{ payment.payment_date }}</td>
                        <td>{{ payment.expected_amount }} XAF</td>
                        <td>{{ payment.paid_amount }} XAF</td>
                        <td>{{ payment.debt_amount }} XAF</td>
                        <td>
                            {% if payment.status == 'paid' %}
                                <span class="status-paid">{{ payment.get_status_display }}</span>
                            {% elif payment.status == 'partial' %}
                                <span class="status-partial">{{ payment.get_status_display }}</span>
                            {% else %}
                                <span class="status-pending">{{ payment.get_status_display }}</span>
                            {% endif %}
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <p class="empty">No hay pagos registrados.</p>
        {% endif %}
    </div>
</details>

<details class="premium-section">
    <summary>Últimos daños</summary>
    <div class="section-body">
        {% if recent_damages %}
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>Taxi</th>
                        <th>Conductor</th>
                        <th>Daño</th>
                        <th>Fecha</th>
                        <th>Costo estimado</th>
                        <th>Estado</th>
                    </tr>
                </thead>
                <tbody>
                    {% for damage in recent_damages %}
                    <tr>
                        <td>{{ damage.vehicle.plate_number }}</td>
                        <td>{{ damage.driver.full_name }}</td>
                        <td>{{ damage.title }}</td>
                        <td>{{ damage.damage_date }}</td>
                        <td>{{ damage.estimated_cost }} XAF</td>
                        <td><span class="status-pending">{{ damage.get_status_display }}</span></td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
            <p class="empty">No hay daños registrados.</p>
        {% endif %}
    </div>
</details>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
document.addEventListener("DOMContentLoaded", function () {
    const money = value => Number(String(value).replace(",", ".")) || 0;

    const paid = money("{{ total_paid|default:0 }}");
    const debt = money("{{ total_debt|default:0 }}");
    const damages = money("{{ damages_cost|default:0 }}");
    const net = money("{{ net_profit|default:0 }}");

    const activeVehicles = money("{{ active_vehicles|default:0 }}");
    const activeDrivers = money("{{ active_drivers|default:0 }}");
    const damagesPending = money("{{ damages_pending|default:0 }}");

    document.querySelectorAll(".counter").forEach(el => {
        const target = money(el.dataset.value);
        let current = 0;
        const step = Math.max(target / 40, 1);

        const timer = setInterval(() => {
            current += step;
            if (current >= target) {
                current = target;
                clearInterval(timer);
            }
            el.textContent = Math.round(current).toLocaleString("es-ES") + " XAF";
        }, 22);
    });

    const alertBox = document.getElementById("liveAlert");
    if (debt > 0 || damagesPending > 0) {
        alertBox.classList.add("show");
    }

    const financeChart = document.getElementById("financeChart");
    if (financeChart) {
        new Chart(financeChart, {
            type: "bar",
            data: {
                labels: ["Cobrado", "Deuda", "Daños", "Neto"],
                datasets: [{
                    label: "XAF",
                    data: [paid, debt, damages, net],
                    borderWidth: 1,
                    borderRadius: 12
                }]
            },
            options: {
                responsive: true,
                maintainAspectRatio: false,
                plugins: { legend: { display: false } },
                scales: {
                    y: { beginAtZero: true }
                }
            }
        });
    }

    const operationChart = document.getElementById("operationChart");
    if (operationChart) {
        new Chart(operationChart, {
            type: "doughnut",
            data: {
                labels: ["Taxis activos", "Conductores activos", "Daños pendientes"],
                datasets: [{
                    data: [activeVehicles, activeDrivers, damagesPending],
                    borderWidth: 2
                }]
            },
            options: {
                responsive: true,
                maintainAspectRatio: false
            }
        });
    }

    setInterval(() => {
        const now = new Date().toLocaleTimeString("es-ES");
        console.log("TaxiGE dashboard activo:", now);
    }, 30000);
});
</script>

{% endblock %}
'''

template_path.write_text(template, encoding="utf-8")
print("OK: Dashboard avanzado aplicado correctamente")
