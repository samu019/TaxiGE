from pathlib import Path
import re

p = Path("dashboard/views.py")
text = p.read_text(encoding="utf-8-sig").replace("\ufeff", "")

# Eliminar versiones anteriores si existen
text = re.sub(r"\n@login_required\s*def dashboard_export_excel\(request\):.*?(?=\n@login_required\s*def |\Z)", "\n", text, flags=re.S)
text = re.sub(r"\n@login_required\s*def dashboard_export_pdf\(request\):.*?(?=\n@login_required\s*def |\Z)", "\n", text, flags=re.S)

new_code = r'''

@login_required
def dashboard_export_excel(request):
    from django.http import HttpResponse
    from django.db.models import Sum
    from decimal import Decimal
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from payments.models import DriverPayment
    from damages.models import VehicleDamage
    from companies.models import Company
    from vehicles.models import Vehicle
    from drivers.models import Driver

    user = request.user

    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)
    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost
    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    wb = Workbook()

    dark = "0F172A"
    yellow = "FACC15"
    green = "22C55E"
    gray = "E2E8F0"

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ======================
    # HOJA RESUMEN
    # ======================
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "Reporte Ejecutivo TaxiGE"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Usuario: {user.username}"
    ws["A3"] = "Sistema: TaxiGE Platform"

    resumen = [
        ["Indicador", "Valor"],
        ["Ingresos cobrados", float(total_paid)],
        ["Deuda pendiente", float(total_debt)],
        ["Costo daños", float(damages_cost)],
        ["Beneficio neto", float(net_profit)],
        ["Ratio de cobro", f"{collection_rate}%"],
        ["Empresas", companies.count()],
        ["Taxis activos", vehicles.filter(status='active').count()],
        ["Conductores activos", drivers.filter(status='active').count()],
        ["Daños pendientes", damages.filter(status='pending').count()],
    ]

    start_row = 5
    for r, row in enumerate(resumen, start=start_row):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r == start_row:
                cell.font = Font(bold=True, color="111827")
                cell.fill = PatternFill("solid", fgColor=yellow)

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 26

    # ======================
    # HOJA PAGOS
    # ======================
    ws2 = wb.create_sheet("Pagos")
    headers = ["Conductor", "Taxi", "Fecha", "Esperado", "Pagado", "Deuda", "Estado"]

    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r, pay in enumerate(payments.order_by("-payment_date", "-created_at"), 2):
        row = [
            getattr(pay.driver, "full_name", ""),
            getattr(pay.vehicle, "plate_number", ""),
            str(pay.payment_date),
            float(pay.expected_amount),
            float(pay.paid_amount),
            float(pay.debt_amount),
            pay.get_status_display(),
        ]
        for c, value in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    # ======================
    # HOJA DAÑOS
    # ======================
    ws3 = wb.create_sheet("Daños")
    headers = ["Taxi", "Conductor", "Daño", "Fecha", "Costo estimado", "Estado"]

    for c, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for r, dmg in enumerate(damages.order_by("-damage_date", "-created_at"), 2):
        row = [
            getattr(dmg.vehicle, "plate_number", ""),
            getattr(dmg.driver, "full_name", ""),
            dmg.title,
            str(dmg.damage_date),
            float(dmg.estimated_cost),
            dmg.get_status_display(),
        ]
        for c, value in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 24

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="taxige_reporte_ejecutivo.xlsx"'
    wb.save(response)
    return response


@login_required
def dashboard_export_pdf(request):
    from django.http import HttpResponse
    from django.db.models import Sum
    from decimal import Decimal
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from payments.models import DriverPayment
    from damages.models import VehicleDamage
    from companies.models import Company
    from vehicles.models import Vehicle
    from drivers.models import Driver

    user = request.user

    payments = DriverPayment.objects.filter(company__owner=user)
    damages = VehicleDamage.objects.filter(company__owner=user)
    companies = Company.objects.filter(owner=user)
    vehicles = Vehicle.objects.filter(company__owner=user)
    drivers = Driver.objects.filter(company__owner=user)

    total_expected = payments.aggregate(total=Sum('expected_amount'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('paid_amount'))['total'] or Decimal('0')
    total_debt = total_expected - total_paid
    damages_cost = damages.aggregate(total=Sum('estimated_cost'))['total'] or Decimal('0')
    net_profit = total_paid - damages_cost
    collection_rate = round((total_paid / total_expected) * 100, 2) if total_expected > 0 else Decimal('0')

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="taxige_reporte_ejecutivo.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("<b>Reporte Ejecutivo TaxiGE</b>", styles["Title"]))
    story.append(Paragraph(f"Usuario: {user.username}", styles["Normal"]))
    story.append(Spacer(1, 14))

    resumen = [
        ["Indicador", "Valor"],
        ["Ingresos cobrados", f"{total_paid} XAF"],
        ["Deuda pendiente", f"{total_debt} XAF"],
        ["Costo daños", f"{damages_cost} XAF"],
        ["Beneficio neto", f"{net_profit} XAF"],
        ["Ratio de cobro", f"{collection_rate}%"],
        ["Empresas", companies.count()],
        ["Taxis activos", vehicles.filter(status='active').count()],
        ["Conductores activos", drivers.filter(status='active').count()],
        ["Daños pendientes", damages.filter(status='pending').count()],
    ]

    table = Table(resumen, colWidths=[220, 220])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Últimos pagos</b>", styles["Heading2"]))
    pagos_data = [["Conductor", "Taxi", "Fecha", "Pagado", "Estado"]]
    for pay in payments.order_by("-payment_date", "-created_at")[:8]:
        pagos_data.append([
            getattr(pay.driver, "full_name", ""),
            getattr(pay.vehicle, "plate_number", ""),
            str(pay.payment_date),
            f"{pay.paid_amount} XAF",
            pay.get_status_display(),
        ])

    pagos_table = Table(pagos_data, colWidths=[120, 80, 85, 95, 95])
    pagos_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#facc15")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(pagos_table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>Últimos daños</b>", styles["Heading2"]))
    damages_data = [["Taxi", "Conductor", "Daño", "Costo", "Estado"]]
    for dmg in damages.order_by("-damage_date", "-created_at")[:8]:
        damages_data.append([
            getattr(dmg.vehicle, "plate_number", ""),
            getattr(dmg.driver, "full_name", ""),
            dmg.title,
            f"{dmg.estimated_cost} XAF",
            dmg.get_status_display(),
        ])

    damages_table = Table(damages_data, colWidths=[80, 120, 130, 90, 75])
    damages_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(damages_table)

    doc.build(story)
    return response
'''

text = text.rstrip() + "\n" + new_code + "\n"

p.write_text(text, encoding="utf-8")
print("OK: exportaciones PDF/Excel premium aplicadas")
